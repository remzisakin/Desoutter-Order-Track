import os
import uuid
from datetime import datetime
from typing import List, Optional, Dict, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXCEL_PATH = os.path.join("data", "Desoutter Order Track.xlsx")
RECORDS_SHEET = "Records"
DATA_SHEET = "Data"  # Salesman & Region yönetimi

# Excel başlıklarının hedef sıralaması
RECORD_FIELD_ORDER = [
    ("date_of_request", "Date of Request"),
    ("salesman", "Sales Person"),
    ("region", "Sales Person Region"),
    ("customer_name", "Customer Name"),
    ("customer_po_no", "Customer PO No"),
    ("salesforce_reference", "Salesforce Reference"),
    ("so_no", "SO No"),
    ("definition", "Definition"),
    ("amount_eur", "Amount (EUR)"),
    ("total_discount_pct", "Total Discount (%)"),
    ("cpi_eur", "CPI (EUR)"),
    ("cps_eur", "CPS (EUR)"),
    ("date_of_delivery", "Date of Delivery"),
    ("date_of_invoice", "Date of Invoice"),
    ("note", "Note"),
    ("record_id", "Record ID"),
    ("created_at", "Created At"),
    ("updated_at", "Updated At"),
]

RECORD_COLUMNS = [column for _, column in RECORD_FIELD_ORDER]
FIELD_TO_COLUMN = {field: column for field, column in RECORD_FIELD_ORDER}
COLUMN_TO_FIELD = {column: field for field, column in RECORD_FIELD_ORDER}

DATA_COLUMNS = ["Sales Person", "Region"]  # Region: CPI Northern / CPI Southern / Unassigned


def get_excel_path() -> str:
    """Excel dosya yolunu döndürür ve yoksa oluşturur."""
    _ensure_file_structure()
    return EXCEL_PATH


def ensure_workbook_format():
    """Excel dosyasındaki sayfaların başlık/biçimlendirmelerini uygular."""
    _ensure_file_structure()
    _apply_records_formatting()
    _apply_data_formatting()


def _ensure_file_structure():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    if not os.path.exists(EXCEL_PATH):
        with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl") as writer:
            pd.DataFrame(columns=RECORD_COLUMNS).to_excel(
                writer, sheet_name=RECORDS_SHEET, index=False
            )
            pd.DataFrame(columns=DATA_COLUMNS).to_excel(
                writer, sheet_name=DATA_SHEET, index=False
            )
        return

    # Eksik sayfaları ekle
    with pd.ExcelWriter(
        EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="overlay"
    ) as writer:
        workbook = writer.book
        if RECORDS_SHEET not in workbook.sheetnames:
            pd.DataFrame(columns=RECORD_COLUMNS).to_excel(
                writer, sheet_name=RECORDS_SHEET, index=False
            )
        if DATA_SHEET not in workbook.sheetnames:
            pd.DataFrame(columns=DATA_COLUMNS).to_excel(
                writer, sheet_name=DATA_SHEET, index=False
            )


def _read_df(sheet: str) -> pd.DataFrame:
    _ensure_file_structure()
    try:
        return pd.read_excel(EXCEL_PATH, sheet_name=sheet)
    except Exception:
        return pd.DataFrame()


def _normalize_record_df(df: pd.DataFrame) -> pd.DataFrame:
    legacy_map = {
        "record_id": FIELD_TO_COLUMN["record_id"],
        "SalesMan": FIELD_TO_COLUMN["salesman"],
        "Region": FIELD_TO_COLUMN["region"],
        "SalesForce Reference": FIELD_TO_COLUMN["salesforce_reference"],
        "Defination": FIELD_TO_COLUMN["definition"],
        "created_at": FIELD_TO_COLUMN["created_at"],
        "updated_at": FIELD_TO_COLUMN["updated_at"],
    }
    df = df.rename(columns=legacy_map)

    for column in RECORD_COLUMNS:
        if column not in df.columns:
            df[column] = None

    df = df[RECORD_COLUMNS]

    # Tarih alanları
    date_columns = [
        FIELD_TO_COLUMN["date_of_request"],
        FIELD_TO_COLUMN["date_of_delivery"],
        FIELD_TO_COLUMN["date_of_invoice"],
        FIELD_TO_COLUMN["created_at"],
        FIELD_TO_COLUMN["updated_at"],
    ]
    for column in date_columns:
        df[column] = pd.to_datetime(df[column], errors="coerce")

    # Numerik alanlar
    numeric_columns = [
        FIELD_TO_COLUMN["amount_eur"],
        FIELD_TO_COLUMN["total_discount_pct"],
        FIELD_TO_COLUMN["cpi_eur"],
        FIELD_TO_COLUMN["cps_eur"],
    ]
    for column in numeric_columns:
        df[column] = pd.to_numeric(df[column], errors="coerce")

    # Metin alanları
    text_columns = [
        FIELD_TO_COLUMN["salesman"],
        FIELD_TO_COLUMN["region"],
        FIELD_TO_COLUMN["customer_name"],
        FIELD_TO_COLUMN["customer_po_no"],
        FIELD_TO_COLUMN["salesforce_reference"],
        FIELD_TO_COLUMN["so_no"],
        FIELD_TO_COLUMN["definition"],
        FIELD_TO_COLUMN["note"],
        FIELD_TO_COLUMN["record_id"],
    ]
    for column in text_columns:
        df[column] = df[column].fillna("").astype(str).str.strip()

    return df


def _normalize_sales_df(df: pd.DataFrame) -> pd.DataFrame:
    legacy_map = {"SalesMan": DATA_COLUMNS[0]}
    df = df.rename(columns=legacy_map)

    for column in DATA_COLUMNS:
        if column not in df.columns:
            df[column] = None

    df = df[DATA_COLUMNS]
    df[DATA_COLUMNS[0]] = df[DATA_COLUMNS[0]].fillna("").astype(str).str.strip()
    df[DATA_COLUMNS[1]] = df[DATA_COLUMNS[1]].fillna("Unassigned").astype(str).str.strip()
    return df


def _write_df(sheet: str, df: pd.DataFrame):
    _ensure_file_structure()
    if sheet == RECORDS_SHEET:
        df = _normalize_record_df(df)
    elif sheet == DATA_SHEET:
        df = _normalize_sales_df(df)

    with pd.ExcelWriter(
        EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        df.to_excel(writer, sheet_name=sheet, index=False)

    if sheet == RECORDS_SHEET:
        _apply_records_formatting()
    elif sheet == DATA_SHEET:
        _apply_data_formatting()


def _apply_records_formatting():
    if not os.path.exists(EXCEL_PATH):
        return
    try:
        workbook = load_workbook(EXCEL_PATH)
    except Exception:
        return
    if RECORDS_SHEET not in workbook.sheetnames:
        workbook.save(EXCEL_PATH)
        return

    ws = workbook[RECORDS_SHEET]

    header_fill = PatternFill(fill_type="solid", fgColor="6AA84F")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    column_widths = {
        FIELD_TO_COLUMN["date_of_request"]: 16,
        FIELD_TO_COLUMN["salesman"]: 20,
        FIELD_TO_COLUMN["region"]: 18,
        FIELD_TO_COLUMN["customer_name"]: 28,
        FIELD_TO_COLUMN["customer_po_no"]: 22,
        FIELD_TO_COLUMN["salesforce_reference"]: 24,
        FIELD_TO_COLUMN["so_no"]: 18,
        FIELD_TO_COLUMN["definition"]: 36,
        FIELD_TO_COLUMN["amount_eur"]: 16,
        FIELD_TO_COLUMN["total_discount_pct"]: 18,
        FIELD_TO_COLUMN["cpi_eur"]: 16,
        FIELD_TO_COLUMN["cps_eur"]: 16,
        FIELD_TO_COLUMN["date_of_delivery"]: 16,
        FIELD_TO_COLUMN["date_of_invoice"]: 16,
        FIELD_TO_COLUMN["note"]: 28,
        FIELD_TO_COLUMN["record_id"]: 38,
        FIELD_TO_COLUMN["created_at"]: 22,
        FIELD_TO_COLUMN["updated_at"]: 22,
    }
    for index, column_name in enumerate(RECORD_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = column_widths.get(column_name, 18)

    currency_columns = {
        FIELD_TO_COLUMN["amount_eur"],
        FIELD_TO_COLUMN["cpi_eur"],
        FIELD_TO_COLUMN["cps_eur"],
    }
    discount_column = FIELD_TO_COLUMN["total_discount_pct"]
    date_columns = {
        FIELD_TO_COLUMN["date_of_request"],
        FIELD_TO_COLUMN["date_of_delivery"],
        FIELD_TO_COLUMN["date_of_invoice"],
    }
    datetime_columns = {
        FIELD_TO_COLUMN["created_at"],
        FIELD_TO_COLUMN["updated_at"],
    }

    for idx, column_name in enumerate(RECORD_COLUMNS, start=1):
        column_letter = get_column_letter(idx)
        data_cells = ws[column_letter][1:]
        if column_name in currency_columns:
            for cell in data_cells:
                cell.number_format = "#,##0.00"
        elif column_name == discount_column:
            for cell in data_cells:
                cell.number_format = "0.00"
        elif column_name in date_columns:
            for cell in data_cells:
                cell.number_format = "yyyy-mm-dd"
        elif column_name in datetime_columns:
            for cell in data_cells:
                cell.number_format = "yyyy-mm-dd hh:mm"

    workbook.save(EXCEL_PATH)


def _apply_data_formatting():
    if not os.path.exists(EXCEL_PATH):
        return
    try:
        workbook = load_workbook(EXCEL_PATH)
    except Exception:
        return
    if DATA_SHEET not in workbook.sheetnames:
        workbook.save(EXCEL_PATH)
        return

    ws = workbook[DATA_SHEET]
    header_fill = PatternFill(fill_type="solid", fgColor="9FC5E8")
    header_font = Font(color="000000", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    ws.column_dimensions[get_column_letter(1)].width = 24
    ws.column_dimensions[get_column_letter(2)].width = 18

    workbook.save(EXCEL_PATH)


# ------------------ Salesman Data ------------------

def list_salesmen() -> pd.DataFrame:
    df = _read_df(DATA_SHEET)
    if df.empty:
        df = pd.DataFrame(columns=DATA_COLUMNS)
    return _normalize_sales_df(df)


def upsert_salesman(name: str, region: str):
    df = list_salesmen()
    mask = df[DATA_COLUMNS[0]].str.lower() == name.strip().lower()
    if mask.any():
        df.loc[mask, DATA_COLUMNS[1]] = region
    else:
        df = pd.concat(
            [
                df,
                pd.DataFrame(
                    [{DATA_COLUMNS[0]: name.strip(), DATA_COLUMNS[1]: region}]
                ),
            ],
            ignore_index=True,
        )
    _write_df(DATA_SHEET, df)


def bulk_set_salesmen(items: List[Dict[str, str]]):
    rows = []
    for it in items:
        rows.append(
            {
                DATA_COLUMNS[0]: it["name"],
                DATA_COLUMNS[1]: it.get("region", "Unassigned"),
            }
        )
    df = pd.DataFrame(rows, columns=DATA_COLUMNS)
    _write_df(DATA_SHEET, df)


# ------------------ Records ------------------

def list_records() -> pd.DataFrame:
    df = _read_df(RECORDS_SHEET)
    if df.empty:
        df = pd.DataFrame(columns=RECORD_COLUMNS)
    return _normalize_record_df(df)


def _infer_region_for_salesman(salesman: str) -> str:
    df = list_salesmen()
    if df.empty:
        return "Unassigned"
    mask = df[DATA_COLUMNS[0]].str.lower() == str(salesman).strip().lower()
    if mask.any():
        region = df.loc[mask, DATA_COLUMNS[1]].iloc[0]
        return region or "Unassigned"
    return "Unassigned"


def _coerce_date(value: Optional[Any]) -> Optional[datetime]:
    if value in (None, "", "nan", "NaT"):
        return None
    try:
        return pd.to_datetime(value, errors="coerce")
    except Exception:
        return None


def create_record(payload: Dict[str, Any]) -> Dict[str, Any]:
    df = list_records()

    record_id = str(uuid.uuid4())
    now = datetime.now()

    amount = float(payload["amount_eur"])
    cps = float(payload.get("cps_eur", 0.0) or 0.0)
    cpi = amount - cps if cps else amount

    region = _infer_region_for_salesman(payload["salesman"])

    row = {column: None for column in RECORD_COLUMNS}
    row[FIELD_TO_COLUMN["record_id"]] = record_id
    row[FIELD_TO_COLUMN["date_of_request"]] = _coerce_date(payload["date_of_request"])
    row[FIELD_TO_COLUMN["salesman"]] = payload["salesman"]
    row[FIELD_TO_COLUMN["region"]] = region
    row[FIELD_TO_COLUMN["customer_name"]] = payload["customer_name"]
    row[FIELD_TO_COLUMN["customer_po_no"]] = payload["customer_po_no"]
    row[FIELD_TO_COLUMN["salesforce_reference"]] = payload["salesforce_reference"]
    row[FIELD_TO_COLUMN["so_no"]] = payload["so_no"]
    row[FIELD_TO_COLUMN["definition"]] = payload.get("definition", "")
    row[FIELD_TO_COLUMN["amount_eur"]] = amount
    row[FIELD_TO_COLUMN["total_discount_pct"]] = float(payload["total_discount_pct"])
    row[FIELD_TO_COLUMN["cpi_eur"]] = cpi
    row[FIELD_TO_COLUMN["cps_eur"]] = cps
    row[FIELD_TO_COLUMN["date_of_delivery"]] = _coerce_date(
        payload.get("date_of_delivery")
    )
    row[FIELD_TO_COLUMN["date_of_invoice"]] = _coerce_date(
        payload.get("date_of_invoice")
    )
    row[FIELD_TO_COLUMN["note"]] = payload.get("note", "")
    row[FIELD_TO_COLUMN["created_at"]] = now
    row[FIELD_TO_COLUMN["updated_at"]] = now

    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    _write_df(RECORDS_SHEET, df)
    return row


def find_record(
    so_no: Optional[str] = None, customer_po_no: Optional[str] = None
) -> Optional[Dict[str, Any]]:
    df = list_records()
    if df.empty:
        return None

    so_col = FIELD_TO_COLUMN["so_no"]
    po_col = FIELD_TO_COLUMN["customer_po_no"]

    result = pd.DataFrame()
    if so_no:
        result = df[
            df[so_col].astype(str).str.lower() == so_no.strip().lower()
        ]
    elif customer_po_no:
        result = df[
            df[po_col].astype(str).str.lower() == customer_po_no.strip().lower()
        ]

    if result.empty:
        return None
    return result.iloc[0].to_dict()


def update_record(record_id: str, payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    df = list_records()
    if df.empty:
        return None

    id_col = FIELD_TO_COLUMN["record_id"]
    mask = df[id_col] == record_id
    if not mask.any():
        return None

    amount = float(payload["amount_eur"])
    cps = float(payload.get("cps_eur", 0.0) or 0.0)
    cpi = amount - cps if cps else amount
    region = _infer_region_for_salesman(payload["salesman"])

    df.loc[mask, FIELD_TO_COLUMN["date_of_request"]] = _coerce_date(
        payload["date_of_request"]
    )
    df.loc[mask, FIELD_TO_COLUMN["salesman"]] = payload["salesman"]
    df.loc[mask, FIELD_TO_COLUMN["region"]] = region
    df.loc[mask, FIELD_TO_COLUMN["customer_name"]] = payload["customer_name"]
    df.loc[mask, FIELD_TO_COLUMN["customer_po_no"]] = payload["customer_po_no"]
    df.loc[mask, FIELD_TO_COLUMN["salesforce_reference"]] = payload[
        "salesforce_reference"
    ]
    df.loc[mask, FIELD_TO_COLUMN["so_no"]] = payload["so_no"]
    df.loc[mask, FIELD_TO_COLUMN["definition"]] = payload.get("definition", "")
    df.loc[mask, FIELD_TO_COLUMN["amount_eur"]] = amount
    df.loc[mask, FIELD_TO_COLUMN["total_discount_pct"]] = float(
        payload["total_discount_pct"]
    )
    df.loc[mask, FIELD_TO_COLUMN["cpi_eur"]] = cpi
    df.loc[mask, FIELD_TO_COLUMN["cps_eur"]] = cps
    df.loc[mask, FIELD_TO_COLUMN["date_of_delivery"]] = _coerce_date(
        payload.get("date_of_delivery")
    )
    df.loc[mask, FIELD_TO_COLUMN["date_of_invoice"]] = _coerce_date(
        payload.get("date_of_invoice")
    )
    df.loc[mask, FIELD_TO_COLUMN["note"]] = payload.get("note", "")
    df.loc[mask, FIELD_TO_COLUMN["updated_at"]] = datetime.now()

    _write_df(RECORDS_SHEET, df)
    return df.loc[mask].iloc[0].to_dict()


# ------------------ Reports ------------------

def report_frames() -> Dict[str, pd.DataFrame]:
    df = list_records()
    if df.empty:
        empty_currency = pd.DataFrame(
            columns=[
                FIELD_TO_COLUMN["region"],
                FIELD_TO_COLUMN["amount_eur"],
                FIELD_TO_COLUMN["cpi_eur"],
                FIELD_TO_COLUMN["cps_eur"],
            ]
        )
        return {
            "by_region": empty_currency,
            "or_by_year": pd.DataFrame(columns=["Year", "OR (EUR)"]),
            "oi_by_year": pd.DataFrame(columns=["Year", "OI (EUR)"]),
            "cpi_vs_cps": pd.DataFrame(columns=["Metric", "EUR"]),
        }

    df = df.copy()
    amount_col = FIELD_TO_COLUMN["amount_eur"]
    cpi_col = FIELD_TO_COLUMN["cpi_eur"]
    cps_col = FIELD_TO_COLUMN["cps_eur"]
    region_col = FIELD_TO_COLUMN["region"]
    request_col = FIELD_TO_COLUMN["date_of_request"]
    invoice_col = FIELD_TO_COLUMN["date_of_invoice"]

    for column in [amount_col, cpi_col, cps_col]:
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0.0)

    df["Year_OR"] = pd.to_datetime(df[request_col], errors="coerce").dt.year
    df["Year_OI"] = pd.to_datetime(df[invoice_col], errors="coerce").dt.year

    by_region = (
        df.groupby(region_col)[[amount_col, cpi_col, cps_col]]
        .sum()
        .reset_index()
        .rename(
            columns={
                region_col: "Region",
                amount_col: "Amount (EUR)",
                cpi_col: "CPI (EUR)",
                cps_col: "CPS (EUR)",
            }
        )
    )

    or_by_year = (
        df.groupby("Year_OR")[[amount_col]]
        .sum()
        .reset_index()
        .rename(columns={"Year_OR": "Year", amount_col: "OR (EUR)"})
        .dropna(subset=["Year"])
    )

    filtered = df[df[invoice_col].notna()].copy()
    oi_by_year = (
        filtered.groupby("Year_OI")[[cpi_col, cps_col]]
        .sum()
        .sum(axis=1)
        .reset_index(name="OI (EUR)")
        .rename(columns={"Year_OI": "Year"})
    )

    cpi_vs_cps = pd.DataFrame(
        [
            {"Metric": "CPI (EUR)", "EUR": df[cpi_col].sum()},
            {"Metric": "CPS (EUR)", "EUR": df[cps_col].sum()},
        ]
    )

    return {
        "by_region": by_region,
        "or_by_year": or_by_year,
        "oi_by_year": oi_by_year,
        "cpi_vs_cps": cpi_vs_cps,
    }
